import openpyxl

class Excel(object):
    height_pixel_per_cell=22
    width_pixel_per_cell=100


    def __init__(self,data_matrix,template=None,start_cell='A1'):
        self.table=data_matrix
        self.binded_xlsx_fname=template
        self.binded_xlsx=openpyxl.Workbook() if template==None else openpyxl.load_workbook(template)
        self.active_ws=self.binded_xlsx.active
        self.current_start_cell=start_cell
        self.data_matrix_poslogs=tuple()
        if template!=None:
            sheet_num=len(self.binded_xlsx.worksheets)
            for i in range(0,sheet_num):
                self.active_ws=self.binded_xlsx.worksheets[i]
                datatype=self.active_ws[self.current_start_cell].value
                if datatype=='Association Matrix':
                    chart=TwoDimChart(self)
                    chart.draw_default_chart()
                elif datatype=='List Table':
                    chart=OneDimChart(self)
                    chart.draw_default_chart()
            self.binded_xlsx.create_sheet(sheet_num)
            self.active_ws=self.binded_xlsx.worksheets[sheet_num]



    def _write_table_into_excel(self,table,merge_tuple=None):
        column_start=int(openpyxl.cell.column_index_from_string(self.current_start_cell[0]))
        rowid=int(self.current_start_cell[1:])
        #get row id and column id of a specific cell which is the top left of data matrix

        if merge_tuple is not None:
            for pos_info in merge_tuple:
                self.active_ws.merge_cells(pos_info)

        for row in table:
            columnid=column_start
            for element in row:
                cell=self.active_ws.cell(row=rowid,column=columnid)
                cell.value=element
                columnid+=1
            rowid+=1

        #fill in data in cells



    def set_current_start_cell(self,cellpos):
        self.current_start_cell=cellpos

    def set_table_header_style(self,cellpos):
        pass

    def load_data_matrix(self,data_matrix):
        self.table=data_matrix

    def save_xlsx(self,output_fname=None):
        #write self.table into wb.ws object
        if output_fname is None:
            output_fname='default_output.xlsx'
        self.binded_xlsx.save(output_fname)
    def _gen_reference(self,label_tuple,value_tuple):
        return openpyxl.charts.Reference(self.active_ws,label_tuple,value_tuple)

    def _gen_series(self,values,title=None,labels=None):
        return openpyxl.charts.Series(values,title=title,labels=labels)

    def add_formula(self,cellpos,str_formula):
        self.active_ws[cellpos]=str_formula


    def add_chart(self,label_nested_tuple,value_nested_tuple,pos_tuple,size_tuple):
        #XXXX_nested_tuple:('series name',(1,1),(10,1))
        #pos_tuple=(top,left)  size_tuple=(height,width)
        labels=self._gen_reference(
                     label_nested_tuple[1],
                     label_nested_tuple[2]
                )
        values=self._gen_reference(
                     value_nested_tuple[1],
                     value_nested_tuple[2]
               )

        series=self._gen_series(
                     values,
                     labels=labels,
                     title=label_nested_tuple[0]
               )
        chart=openpyxl.charts.BarChart()
        chart.append(series)
        chart.drawing.top=pos_tuple[0]
        chart.drawing.left=pos_tuple[1]
        chart.drawing.height=size_tuple[0]
        chart.drawing.width=size_tuple[1]
        chart.width=.6
        chart.height=.7

        self.active_ws.add_chart(chart)


    def get_column_letter(self,col_idx):
        return openpyxl.cell.get_column_letter(col_idx)



    def get_highest_row(self):
        return self.active_ws.get_highest_row()

    @property
    def max_row(self):
        return self.get_highest_row()

    def get_highest_column(self):
        return self.active_ws.get_highest_column()

    @property
    def max_col(self):
        return self.get_highest_column()

    def get_current_start_row(self):
        return int(self.current_start_cell[1:])

    @property
    def min_row(self):
        return self.get_current_start_row()

    def get_current_start_column(self):
        return int(openpyxl.cell.column_index_from_string(self.current_start_cell[0]))

    @property
    def min_col(self):
        return self.get_current_start_column()

    def get_data_matrix_size(self):
        return len(self.table)

class CellOffset(object):
    def __init__(self):
        self._row_offset=int()
        self._col_offset=int()

    @property
    def row(self):
        return self._row_offset

    @row.setter
    def row(self,value):
        self._row_offset=value

    @property
    def col(self):
        return self._col_offset

    @col.setter
    def col(self,value):
        self._col_offset=value


class TwoDimChart(object):
    def __init__(self,boshExcelobject):
        self.excel=boshExcelobject

    def draw_default_piechart(self,series,top_offset,left_offset,size):
        piechart=openpyxl.charts.PieChart()
        piechart.drawing.top=top_offset
        piechart.drawing.left=left_offset
        piechart.drawing.height=size
        piechart.drawing.width=size
        piechart.width=.6
        piechart.height=.6

        piechart.append(series)
        self.excel.active_ws.add_chart(piechart)


    def draw_default_chart(self):
        datastart=CellOffset()
        dataend=CellOffset()
        datastart.row=2
        datastart.col=1
        dataend.row=-1
        dataend.col=-1
        #offsets are used to calculate data square

        col_header_at_row=self.excel.min_row+1
        row_header_at_col=self.excel.min_col
        min_chartwidth=800
        chartheight=300

        col_labels=self.excel._gen_reference(
                       (col_header_at_row,self.excel.min_col+datastart.col),
                       (col_header_at_row,self.excel.max_col+dataend.col)
               )
        row_labels=self.excel._gen_reference(
                       (self.excel.min_row+datastart.row,row_header_at_col),
                       (self.excel.max_row+dataend.row,row_header_at_col)
               )

        chart=openpyxl.charts.BarChart()
        chart.drawing.top=self.excel.height_pixel_per_cell*self.excel.max_row
        chart.drawing.left=0
        chart.drawing.height=chartheight
        width=self.excel.width_pixel_per_cell*self.excel.max_col
        chart.drawing.width=width if width > min_chartwidth else min_chartwidth
        chart.width=.8
        chart.height=.8


        for i in range(self.excel.min_row+datastart.row,self.excel.max_row+dataend.row+1):
            values=self.excel._gen_reference(
                     (i,self.excel.min_col+datastart.col),
                     (i,self.excel.max_col+dataend.col)
            )
            cell=self.excel.active_ws.cell(row=i,column=self.excel.min_col)
            series_title=cell.value
            series=self.excel._gen_series(
                   values,
                   labels=col_labels,
                   title=series_title
            )

            chart.append(series)

        self.excel.active_ws.add_chart(chart)

        colsum_values=self.excel._gen_reference(
                             (self.excel.max_row,self.excel.min_col+datastart.col),
                             (self.excel.max_row,self.excel.max_col+dataend.col)
                      )
        series=self.excel._gen_series(
              colsum_values,
              labels=col_labels
        )


        self.draw_default_piechart(series,chart.drawing.top+chartheight,0,chart.drawing.width/2)

        rowsum_values=self.excel._gen_reference(
                            (self.excel.min_row+datastart.row,self.excel.max_col),
                            (self.excel.max_row+dataend.row,self.excel.max_col)
                      )
        series=self.excel._gen_series(
              rowsum_values,
              labels=row_labels
        )
        self.draw_default_piechart(series,chart.drawing.top+chartheight,chart.drawing.width/2,chart.drawing.width/2)


class OneDimChart(object):
    def __init__(self,boshExcelobject):
        self.excel=boshExcelobject

    def draw_default_piechart(self,series,top_offset,left_offset,size):
        piechart=openpyxl.charts.PieChart()
        piechart.drawing.top=top_offset
        piechart.drawing.left=left_offset
        piechart.drawing.height=size
        piechart.drawing.width=size
        piechart.width=.7
        piechart.height=.7

        piechart.append(series)
        self.excel.active_ws.add_chart(piechart)

    def draw_default_chart(self):
        datastart=CellOffset()
        dataend=CellOffset()
        datastart.row=1
        datastart.col=1
        dataend.row=-1
        dataend.col=0
        #offsets are used to calculate data square

        row_header_at_col=self.excel.min_col
        min_chartwidth=600
        chartheight=300
        row_bound=5


        row_labels=self.excel._gen_reference(
                       (self.excel.min_row+datastart.row,row_header_at_col),
                       (self.excel.max_row+dataend.row,row_header_at_col)
               )

        chart=openpyxl.charts.BarChart()
        chart.drawing.top=0
        chart.drawing.height=chartheight
        chart.drawing.left=self.excel.width_pixel_per_cell*self.excel.max_col
        width=self.excel.width_pixel_per_cell*self.excel.max_row/3
        isfewrows=(self.excel.max_row-datastart.row+dataend.row)<=row_bound
        piechartheight=chartheight if isfewrows  else chartheight+50*(self.excel.max_row/row_bound)
        chart.drawing.width=width if width > min_chartwidth else min_chartwidth


        chart.width=.8
        chart.height=.8

        pie_values=self.excel._gen_reference(
                             (self.excel.min_row+datastart.row,self.excel.max_col+dataend.col),
                             (self.excel.max_row+dataend.row,self.excel.max_col+dataend.col)
                      )
        series=self.excel._gen_series(
              pie_values,
              labels=row_labels
        )

        pietop=0 if isfewrows else chartheight


        self.draw_default_piechart(series,pietop,chart.drawing.left,piechartheight)

        chart.drawing.left=chart.drawing.left+piechartheight if isfewrows else chart.drawing.left

        for i in range(self.excel.min_row+datastart.row,self.excel.max_row+dataend.row+1):
            values=self.excel._gen_reference(
                     (i,self.excel.min_col+datastart.col),
                     (i,self.excel.max_col+dataend.col)
            )
            cell=self.excel.active_ws.cell(row=i,column=self.excel.min_col)
            series_title=cell.value
            series=self.excel._gen_series(
                   values,
                   title=series_title
            )

            chart.append(series)

        self.excel.active_ws.add_chart(chart)



class matrixSheet(Excel):

    def __init__(self,data_matrix,template=None,start_cell='A1'):
        super(matrixSheet,self).__init__(data_matrix,template,start_cell)
        self.chart=TwoDimChart(self)



    def write_twodim_table(self):
        row_header=list(set(record[0] for record in self.table))
        col_header=list(set(record[1] for record in self.table))

        data_start_row_idx=2
        data_start_col_idx=1
        row_dict={j:i for i,j in enumerate(row_header,start=data_start_row_idx)}
        col_dict={j:i for i,j in enumerate(col_header,start=data_start_col_idx)}
        #row header/col header both occupy 1 cell, start from 1

        additional_col=2
        additional_row=3
        #preserve 1 for header, 1 for summation, over each row/col

        new_table=[[0 for i in range(0,len(col_header)+additional_col)] for i in range(0,len(row_header)+additional_row)]


        sum_each_row=[0 for i in range(0,len(row_header)+additional_row)]
        sum_each_col=[0 for i in range(0,len(col_header)+additional_col)]
        for record in self.table:
            row_idx=row_dict[record[0]]
            col_idx=col_dict[record[1]]
            new_table[row_idx][col_idx]=record[2]
            sum_each_row[row_idx]+=record[2]
            sum_each_col[col_idx]+=record[2]

        table_header_pos_at_row=0
        col_header_pos_at_row=1
        row_header_pos_at_col=0
        new_table[table_header_pos_at_row][0]='Association Matrix'


        for i,j in enumerate(col_header,start=data_start_col_idx):
            new_table[col_header_pos_at_row][i]=j
        for i,j in enumerate(row_header,start=data_start_row_idx):
            new_table[i][row_header_pos_at_col]=j

        for i,j in enumerate(sum_each_row):
            new_table[i][len(sum_each_col)-1]=j
        for i,j in enumerate(sum_each_col):
            new_table[len(sum_each_row)-1][i]=j

        for i in [col_header_pos_at_row,len(sum_each_row)-1]:
            for j in [row_header_pos_at_col,len(sum_each_col)-1]:
                new_table[i][j]=None


        merge_str="{}:{}{}".format(
                       self.current_start_cell,
                       self.get_column_letter(len(new_table[0])),
                       self.get_current_start_row()
                  )
        merge_tuple=list()
        merge_tuple.append(merge_str)
        self._write_table_into_excel(new_table,merge_tuple=merge_tuple)
        self.chart.draw_default_chart()






class OneDimSheet(Excel):
    def __init__(self,data_matrix,template=None,start_cell='A1'):
        self.chart=OneDimChart(self)
        super(OneDimSheet,self).__init__(data_matrix,template,start_cell)

    def write_onedim_table(self):
        row_header=list(set(record[0] for record in self.table))
        data_start_at_col=1
        row_dict={j:i for i,j in enumerate(row_header,start=data_start_at_col)}
        additional_row=2
        additional_col=1
        col_num=1
        new_table=[[0 for col in range(0,col_num+additional_col)] for row in range(0,len(row_header)+additional_row)]
        sum_of_col=sum(record[1] for record in self.table)
        table_header_pos_at_row=0
        row_header_pos_at_col=0
        data_pos_at_col=1
        new_table[table_header_pos_at_row][0]='List Table'

        for record in self.table:
            row_id=row_dict[record[0]]
            new_table[row_id][data_pos_at_col]=record[1]
            new_table[row_id][row_header_pos_at_col]=record[0]

        new_table[len(new_table)-1][row_header_pos_at_col]=None
        new_table[len(new_table)-1][data_pos_at_col]=sum_of_col

        merge_str="{}:{}{}".format(
              self.current_start_cell,
                    self.get_column_letter(len(new_table[0])),
                    self.get_current_start_row()
              )
        merge_tuple=list()
        merge_tuple.append(merge_str)
        self._write_table_into_excel(new_table,merge_tuple=merge_tuple)
        self.chart.draw_default_chart()

class listSheet(Excel):

    def __init__(self,data_matrix,template=None,start_cell='A1'):
        super(listSheet,self).__init__(data_matrix,template,start_cell)



    def write_list_of_results(self,table_name=None):
        row_len=len(self.table[0])
        self.table.insert(0,['List of Query Results'])

        merge_str="{}:{}{}".format(
                       self.current_start_cell,
                       self.get_column_letter(row_len),
                       self.get_current_start_row()
                  )
        merge_tuple=list()
        merge_tuple.append(merge_str)

        self._write_table_into_excel(self.table,merge_tuple=merge_tuple)



def Sql_matrix_output(result_table,template=None,output=None):
    matrix_sheet=matrixSheet(result_table,template=template)
    matrix_sheet.write_twodim_table()
    matrix_sheet.save_xlsx(output)

def Sql_list_output(result_table,template=None,output=None):
    list_sheet=listSheet(result_table,template=template)
    list_sheet.write_list_of_results()
    list_sheet.save_xlsx(output)

def Sql_onedim_output(result_table,template=None,output=None):
    onedim_sheet=OneDimSheet(result_table,template=template)
    onedim_sheet.write_onedim_table()
    onedim_sheet.save_xlsx(output)



def Default_list_output(result_table,template=None,output=None):
    list_sheet=listSheet(result_table,template=template)
    list_sheet.write_list_of_results()
    list_sheet.save_xlsx(output)


def Default_output(result_table,template=None,output=None):
    datarowlen=len(result_table[0])
    if(datarowlen==2):
        onedim_sheet=OneDimSheet(result_table,template=template)
        onedim_sheet.write_onedim_table()
        onedim_sheet.save_xlsx(output)
    elif(datarowlen==3):
        matrix_sheet=matrixSheet(result_table,template=template)
        matrix_sheet.write_twodim_table()
        matrix_sheet.save_xlsx(output)
    else:
        list_sheet=listSheet(result_table,template=template)
        list_sheet.write_list_of_results()
        list_sheet.save_xlsx(output)
