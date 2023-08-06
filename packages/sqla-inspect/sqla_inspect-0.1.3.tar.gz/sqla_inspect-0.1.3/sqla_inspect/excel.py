# -*- coding: utf-8 -*-
# * Authors:
#       * Arezki Feth <f.a@majerti.fr>;
#       * Miotte Julien <j.m@majerti.fr>;
#       * TJEBBES Gaston <g.t@majerti.fr>
"""
Excel exportation module
"""

import itertools
import logging
import openpyxl
from openpyxl.styles import (
    Color,
    Font,
)

import cStringIO as StringIO
from string import ascii_uppercase

from sqla_inspect.export import (
    BaseExporter,
    SqlaExporter,
)


log = logging.getLogger(__name__)

# A, B, C, ..., AA, AB, AC, ..., ZZ
ASCII_UPPERCASE = list(ascii_uppercase) + list(
    ''.join(duple)
    for duple in itertools.combinations_with_replacement(ascii_uppercase, 2)
    )


class XlsWriter(object):
    """
    Class providing common tools to write excel files from tabular datas

    Has to be subclassed, the subclass should provide a _datas and a headers
    attribute that contains the datas to render and the headers

        _datas

            list of tuples (each tuple is a row)

        headers

            list of dict containing the label of each column:
                {'label': <a label>}

    """
    title = u"Export"
    def __init__(self):
        self.book = openpyxl.workbook.Workbook()
        self.worksheet = self.book.active
        self.worksheet.title = self.title

    def save_book(self, f_buf=None):
        """
        Return a file buffer containing the resulting xls

        :param obj f_buf: A file buffer supporting the write and seek
        methods
        """
        if f_buf is None:
            f_buf = StringIO.StringIO()
        f_buf.write(openpyxl.writer.excel.save_virtual_workbook(self.book))
        f_buf.seek(0)
        return f_buf

    def set_color(self, cell, color):
        """
        Set the given color to the provided cell

            cell

                A xls cell object

            color

                A openpyxl color var
        """
        cell.style = cell.style.copy(font=Font(color=Color(rgb=color)))

    def format_row(self, row):
        """
        The render method expects rows as lists, here we switch our row format
        from dict to list respecting the order of the headers
        """
        res = []
        headers = getattr(self, 'headers', [])
        for column in headers:
            column_name = column['name']
            value = row.get(column_name, '')
            res.append(value)
        return res

    def render(self, f_buf=None):
        """
        Definitely render the workbook

        :param obj f_buf: A file buffer supporting the write and seek
        methods
        """
        self._render_headers()
        self._render_rows()

        return self.save_book(f_buf)


    def _render_rows(self):
        """
        Render the rows in the current stylesheet
        """
        _datas = getattr(self, '_datas', ())
        for index, row in enumerate(_datas):
            row_number = index + 2
            for col_num, value in enumerate(row):
                cell = self.worksheet.cell(row=row_number, column=col_num + 1)
                cell.value = value

    def _render_headers(self):
        """
        Write the headers row
        """
        headers = getattr(self, 'headers', ())
        for index, col in enumerate(headers):
            # We write the headers
            cell = self.worksheet.cell(row=1, column=index + 1)
            cell.value = col['label']


class SqlaXlsExporter(XlsWriter, SqlaExporter):
    """
    Main class used for exporting datas to the xls format

    Models attributes output can be customized through the info param :

        Column(Integer, infos={'export':
            {'excel':<excel_specific_options>,
             <main_export_options>
            }
        }

    main_export_options and excel_specific_options can be :

        label

            label of the column header

        format

            a function that will be fired on each row to format the output

        related_key

            If the attribute is a relationship, the value of the given attribute
            of the related object will be used to fill the cells

        exclude

            This data will not be inserted in the export if True

    Usage:

        a = SqlaXlsExporter(MyModel)
        for i in MyModel.query().filter(<myfilter>):
            a.add_row(i)
        a.render()
    """
    config_key = 'excel'
    def __init__(self, model):
        XlsWriter.__init__(self)
        SqlaExporter.__init__(self, model)


class XlsExporter(XlsWriter, BaseExporter):
    """
    A main xls exportation tool (without sqlalchemy support)

    writer = MyXlsExporter()
    writer.add_row({'key': u'La valeur de la cellule de la colonne 1'})
    writer.render()
    """
    headers = ()

    def __init__(self):
        XlsWriter.__init__(self)
        BaseExporter.__init__(self)
