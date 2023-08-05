###############################################################################
#
#  Agora Portfolio & Risk Management System
#
#  Description: risk monitoring
#
#  File Created: 01 Jan 2015
#  Author: Carlo Sbraccia
#
###############################################################################

from Agora.LIB_PortfolioFns import LeavesIterator, createUnique, Security
from Agora.LIB_PortfolioIOFns import LoadPositionsFromFile, PortNode
from Agora.LIB_DateFns import DateRange
from Agora.CFG_FundParameters import FUNDFOLDER

import numpy as np
import openpyxl
import sys
import os
import pandas

__all__ = [ ]

TRADES_FILE = os.path.join( FUNDFOLDER, "Utilities Fund", "Live Portfolio", "PMS_LIVE.xlsm" )
SECTORS_FILE = os.path.join( FUNDFOLDER, "Utilities Fund", "Live Portfolio",
                             "CFP Equity Universe", "CFP Equity Universe 201409.xlsx" )

################################################################################
##  loadUniverse
##------------------------------------------------------------------------------
def loadUniverse( target_region = None, target_sector = None ):
    wb = openpyxl.load_workbook( SECTORS_FILE, data_only = True )
    sh = wb.get_sheet_by_name( "Lookup By Ticker" )

    securities = []
    irow = 1
    while True:
        # --- pre-increment row counter
        irow += 1

        sec = sh.cell( row = irow, column = 1 ).value

        if sec is None:
            break

        ccy = str( sh.cell( row = irow, column = 4 ).value )

        if target_region is not None:
            region = str( sh.cell( row = irow, column = 6 ).value )
            if region != target_region: continue

        if target_sector is not None:
            sector = str( sh.cell( row = irow, column = 7 ).value )
            if sector != target_sector: continue

        securities.append( createUnique( str( sec ), ccy, Security ) )

    return securities

################################################################################
##  createIndex
##------------------------------------------------------------------------------
def createIndex( sd, ed, target_region = None, target_sector = None ):
    dfs = []
    for sec in loadUniverse( target_region, target_sector ):
        try:
            df = sec.historicalPricesUSD( sd, ed  )
        except IndexError:
            continue
        dfs.append( df )

    agg = pandas.concat( dfs, axis = 1 )
    ret = np.log( agg.fillna( method = "ffill" ) ).diff().mean( axis = 1 )
    idx = np.cumprod( np.exp( ret ) )
    idx[0] = 1.0

    return idx

################################################################################
##  loadLookupByStrategy
##------------------------------------------------------------------------------
def loadLookupByStrategy( col_id = 2 ):
    wb = openpyxl.load_workbook( SECTORS_FILE, data_only = True )
    sh = wb.get_sheet_by_name( "Lookup By Strategy" )

    lookup = {}

    irow = 1
    while True:
        # --- pre-increment row counter
        irow += 1

        # --- strategy
        strat = sh.cell( row = irow, column = 1 ).value

        if strat is None:
            break
        else:
            strat = str( strat )

        lookup[strat] = str( sh.cell( row = irow, column = col_id ).value )

    return lookup

################################################################################
##  topLevelAnalysis
##------------------------------------------------------------------------------
def topLevelAnalysis( sd, ed, fout = sys.stdout ):
    print( "date,long_mtm,short_mtm,"
           "fund_gross,long_gross,short_gross", file = fout )

    for pd in DateRange( sd, ed, "+1b" ):
        fund = LoadPositionsFromFile( TRADES_FILE, dateCutOff = pd )

        long_mtm  = fund.MTM( pd, "Long" )
        short_mtm = fund.MTM( pd, "Short" )

        fund_gross  = fund.GrossExposure( pd )
        long_gross  = fund.GrossExposure( pd, "Long" )
        short_gross = fund.GrossExposure( pd, "Short" )

        print( "%s,%f,%f,%f,%f,%f" % ( pd,
               long_mtm, short_mtm, fund_gross, long_gross, short_gross ), file = fout )

        if fout != sys.stdout:
            print( "%s,%f,%f,%f,%f,%f" % ( pd,
                   long_mtm, short_mtm, fund_gross, long_gross, short_gross ) )

################################################################################
##  sectorLevelAnalysis
##------------------------------------------------------------------------------
def sectorLevelAnalysis( sd, ed, aggregator, sector, rule = "+1b", fout = sys.stdout ):
    print( "date,sct_long_mtm,sct_short_mtm,",
           "fund_gross,sct_long_gross,sct_short_gross", file = fout )

    for pd in DateRange( sd, ed, rule ):
        fund = LoadPositionsFromFile( TRADES_FILE, dateCutOff = pd )
        sct  = PortNode( name = sector, nodeId = 1 )

        for leaf in LeavesIterator( fund ):
            tag = aggregator( leaf )
            if tag == sector:
                sct.addChild( leaf.clone() )

        long_mtm  = sct.MTM( pd, "Long" )
        short_mtm = sct.MTM( pd, "Short" )

        fund_gross  = fund.GrossExposure( pd )
        long_gross  = sct.GrossExposure( pd, "Long" )
        short_gross = sct.GrossExposure( pd, "Short" )

        print( "%s,%f,%f,%f,%f,%f" % ( pd,
               long_mtm, short_mtm, fund_gross, long_gross, short_gross ), file = fout )

################################################################################
##  aggregatedLevelAnalysis
##------------------------------------------------------------------------------
def aggregatedLevelAnalysis( sd, ed, aggregator, fout = sys.stdout ):
    print( "date,agg_mtm,agg_long_mtm,agg_short_mtm,",
           "agg_gross,agg_long_gross,agg_short_gross", file = fout )

    for pd in DateRange( sd, ed, "+1b" ):

        fund = LoadPositionsFromFile( TRADES_FILE, dateCutOff = pd )
        root = PortNode( name = "root", nodeId = 1 )

        for leaf in LeavesIterator( fund ):
            tag = aggregator( leaf )
            aggNode = root.addChild( PortNode( name = tag ) )
            aggNode.addChild( leaf.clone() )

    return root