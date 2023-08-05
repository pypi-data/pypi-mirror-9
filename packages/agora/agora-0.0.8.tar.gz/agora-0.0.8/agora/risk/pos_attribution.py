###############################################################################
#
#  Agora Portfolio & Risk Management System
#
#  Description: risk monitoring
#
#  File Created: 01 Jan 2015
#  Author: Carlo Sbraccia
#
###############################################################################

import openpyxl
import os

################################################################################
##  loadSectorByTicker
##------------------------------------------------------------------------------
def loadSectorByTicker( sectorsFile ):
    wb = openpyxl.load_workbook( sectorsFile, data_only = True )
    sh = wb.get_sheet_by_name( "Sector By Ticker" )

    sectorByTicker = {}

    irow = 0
    while True:
        # --- pre-increment row counter
        irow += 1

        # --- ticker
        ticker = sh.cell( row = irow, column = 0 ).value

        if ticker is None:
            break
        else:
            ticker = str( ticker ).upper()

        region = str( sh.cell( row = irow, column = 4 ).value )
        sector = str( sh.cell( row = irow, column = 5 ).value )

        sectorByTicker[ticker] = "%s-%s" % ( sector, region )

    return sectorByTicker

if __name__ == "__main__":
    from Snail.LIB_Datatypes import Date
    from Snail.LIB_PortfolioIOFns import LoadPositionsFromFtp
    from Snail.LIB_CurveFns import Interpolate

    import Snail.CFG_FundParameters as cfg
    cfg.setup_config( cfg.__dict__, "main" )

    import math

    sectorsFile = os.path.join( cfg.FUNDFOLDER, "Utilities Fund", "Live Portfolio",
                                "CFP Equity Universe", "CFP Equity Universe 201404.xlsx" )

    mdd = Date( 2014, 3, 31 )
    nav = Interpolate( cfg.NAVCRV, mdd )

    lookup = loadSectorByTicker( sectorsFile )
    lookup["MOZ4 COMDTY"] = "CO2"

    fund = LoadPositionsFromFtp( host = cfg.FTPHOST,
                                 user = cfg.FTPUSER,
                                 passwd = cfg.FTPPASSWD,
                                 date = mdd,
                                 portFilter = cfg.FUNDNAME )

    attribution = {}
    total = 0.0

    for sec, qty in fund.positions.iteritems():
        if sec.isccy:
            continue

        gross = math.fabs( qty*sec.delta( mdd )*sec.underlying.lastUSD( mdd ) )
        total += gross

        name = sec.underlying.name
        name = name.replace( " GR ", " GY " )

        try:        
            sector = lookup[name]
        except KeyError:
            if "INDEX" in name:
                sector = "Index Futures"
            else:
                raise
        
        attribution[sector] = attribution.get( sector, 0.0 ) + gross

    for key, value in attribution.iteritems():
        print key, ",", value / total