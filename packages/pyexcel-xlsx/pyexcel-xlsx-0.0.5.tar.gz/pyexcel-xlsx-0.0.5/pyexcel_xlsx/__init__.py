"""
    pyexcel.ext.xlsx
    ~~~~~~~~~~~~~~~~~~~

    The lower level xlsx file format handler using xlrd/xlwt

    :copyright: (c) 2015 by Onni Software Ltd.
    :license: New BSD License
"""
import sys
import openpyxl
from pyexcel_io import SheetReader, BookReader, SheetWriter, BookWriter
if sys.version_info[0] < 3:
    from StringIO import StringIO
else:
    from io import BytesIO as StringIO


COLUMNS = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
COLUMN_LENGTH = 26


def get_columns(index):
    if index < COLUMN_LENGTH:
        return COLUMNS[index]
    else:
        return (get_columns(int(index // COLUMN_LENGTH) - 1) + COLUMNS[index % COLUMN_LENGTH])


class XLSXSheet(SheetReader):
    """
    xls sheet
    """
    @property
    def name(self):
        return self.native_sheet.title

    def number_of_rows(self):
        """
        Number of rows in the xls sheet
        """
        return self.native_sheet.get_highest_row()

    def number_of_columns(self):
        """
        Number of columns in the xls sheet
        """
        return self.native_sheet.get_highest_column()

    def cell_value(self, row, column):
        """
        Random access to the xls cells
        """
        actual_row = row + 1
        cell_location = "%s%d" % (get_columns(column), actual_row)
        return self.native_sheet.cell(cell_location).value


class XLSXBook(BookReader):
    """
    XLSBook reader

    It reads xls, xlsm, xlsx work book
    """
    def get_sheet(self, nativeSheet):
        return XLSXSheet(nativeSheet)

    def load_from_memory(self, file_content, **keywords):
        return openpyxl.load_workbook(filename=StringIO(file_content),
                                      data_only=True)

    def load_from_file(self, filename, **keywords):
        return openpyxl.load_workbook(filename=filename,
                                      data_only=True)

    def sheet_iterator(self):
        if self.sheet_name is not None:
            sheet = self.native_book.get_sheet_by_name(self.sheet_name)
            if sheet is None:
                raise ValueError("%s cannot be found" % self.sheet_name)
            else:
                return [sheet]
        elif self.sheet_index is not None:
            names = self.native_book.sheetnames
            length = len(names)
            if self.sheet_index < length:
                return [self.native_book.get_sheet_by_name(names[self.sheet_index])]
            else:
                raise IndexError("Index %d of out bound %d" %(self.sheet_index,
                                                              length))
        else:
            return self.native_book


class XLSXSheetWriter(SheetWriter):
    """
    xls, xlsx and xlsm sheet writer
    """
    def set_sheet_name(self, name):
        self.native_sheet.title = name
        self.current_row = 1

    def write_row(self, array):
        """
        write a row into the file
        """
        for i in range(0, len(array)):
            cell_location = "%s%d" % (get_columns(i), self.current_row)
            self.native_sheet.cell(cell_location).value = array[i]
        self.current_row += 1


class XLSXWriter(BookWriter):
    """
    xls, xlsx and xlsm writer
    """
    def __init__(self, file, **keywords):
        BookWriter.__init__(self, file, **keywords)
        self.native_book = openpyxl.Workbook()
        self.current_sheet = 0

    def create_sheet(self, name):
        if self.current_sheet == 0:
            self.current_sheet = 1
            return XLSXSheetWriter(self.native_book,
                                   self.native_book.active, name)
        else:
            return XLSXSheetWriter(self.native_book,
                                   self.native_book.create_sheet(), name)

    def close(self):
        """
        This call actually save the file
        """
        self.native_book.save(filename=self.file)

try:
    from pyexcel.io import READERS
    from pyexcel.io import WRITERS

    READERS.update({
        "xlsm": XLSXBook,
        "xlsx": XLSXBook
    })
    WRITERS.update({
        "xlsm": XLSXWriter,
        "xlsx": XLSXWriter
    })
except:
    # to allow this module to function independently
    pass

__VERSION__ = "0.0.5"
